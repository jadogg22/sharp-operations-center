from __future__ import annotations

from collections import OrderedDict
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.models import CustomerStop

DETAIL_HEADERS = [
    "Order ID",
    "Order date",
    "Delivery date",
    "Bill date",
    "Origin city",
    "Origin state",
    "Origin ZIP",
    "Destination city",
    "Destination state",
    "Destination ZIP",
    "Consignee",
    "Miles",
    "BOL number",
    "Commodity",
    "Weight",
    "Movement",
    "Total pallets",
    "Pallets dropped",
    "Pallets picked up",
    "Linehaul",
    "Fuel surcharge",
    "Extra drops",
    "Extra pickups",
    "Other charges",
    "Other charge total",
    "Total charges",
    "Fuel allocation",
    "Freight allocation",
    "Trailer number",
]

GREEN = "174D37"
DARK_GREEN = "103B2A"
PALE_GREEN = "E8F1EA"
PALE_BLUE = "EEF3F7"
GOLD = "F0C75E"
MUTED = "66716A"
INK = "1F2A23"
WHITE = "FFFFFF"
LINE = "DDE3DE"


def group_stops_by_order(
    stops: list[CustomerStop],
) -> OrderedDict[tuple[str, str], list[CustomerStop]]:
    """Group movement rows in source order for order-level billing sections."""
    # Company ID is part of the key because order IDs may not be globally
    # unique across companies.
    grouped: OrderedDict[tuple[str, str], list[CustomerStop]] = OrderedDict()
    for stop in stops:
        grouped.setdefault((stop.company_id, stop.order_id), []).append(stop)
    return grouped


def invoice_total(stops: list[CustomerStop]) -> float:
    """Calculate the invoice total once per order, not once per movement."""
    # Order-level charges repeat on every movement row, so count only the first
    # row in each order when calculating the invoice total.
    return round(
        sum(order_stops[0].total_charge for order_stops in group_stops_by_order(stops).values()),
        2,
    )


def _excel_date(value):
    return value if value is not None else ""


def _location(city: str, state: str) -> str:
    return ", ".join(part for part in (city.strip(), state.strip()) if part)


def _style_title(sheet, title: str, subtitle: str, last_column: str) -> None:
    sheet.merge_cells(f"A1:{last_column}2")
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Aptos Display", size=22, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=GREEN)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.merge_cells(f"A3:{last_column}3")
    sheet["A3"] = subtitle
    sheet["A3"].font = Font(name="Aptos", size=10, color=MUTED)
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 15


def _build_summary_sheet(
    workbook: Workbook,
    grouped: OrderedDict[tuple[str, str], list[CustomerStop]],
    invoice_number: str,
    bill_date: date | str,
    calculated_total: float,
    expected_total: float | None,
) -> None:
    """Build the customer-facing summary with order and delivery-location rows.

    ``grouped`` supplies one bold parent row per order and its ordered movement
    rows. Charges remain on the parent row so the visible invoice total cannot
    be accidentally multiplied by the number of delivery stops.
    """
    sheet = workbook.active
    sheet.title = "Invoice Summary"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A11"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.outlinePr.summaryBelow = True

    _style_title(
        sheet,
        "CUSTOMER DISTRIBUTION BILLING",
        "Portfolio demonstration  |  Synthetic customers, routes, and financial values",
        "K",
    )

    summary_cards = [
        ("A5:C5", "A6:C7", "Invoice / manifest", invoice_number or "Not provided", "@"),
        ("D5:F5", "D6:F7", "Bill date / period", str(bill_date), "yyyy-mm-dd"),
        ("G5:H5", "G6:H7", "Loads", len(grouped), "#,##0"),
        ("I5:K5", "I6:K7", "Invoice total", calculated_total, '$#,##0.00'),
    ]
    for label_range, value_range, label, value, number_format in summary_cards:
        sheet.merge_cells(label_range)
        sheet.merge_cells(value_range)
        label_cell = sheet[label_range.split(":")[0]]
        value_cell = sheet[value_range.split(":")[0]]
        label_cell.value = label.upper()
        label_cell.font = Font(name="Aptos", size=8, bold=True, color=MUTED)
        label_cell.fill = PatternFill("solid", fgColor=PALE_GREEN)
        label_cell.alignment = Alignment(vertical="center")
        value_cell.value = value
        value_cell.font = Font(name="Aptos Display", size=16, bold=True, color=GREEN)
        value_cell.fill = PatternFill("solid", fgColor=PALE_GREEN)
        value_cell.alignment = Alignment(vertical="center")
        value_cell.number_format = number_format

    sheet.merge_cells("A8:K8")
    verification = sheet["A8"]
    if expected_total is None:
        verification.value = "EXPECTED TOTAL NOT ENTERED - Review the calculated invoice total before sending."
        verification.fill = PatternFill("solid", fgColor="FFF5D6")
        verification.font = Font(name="Aptos", size=9, bold=True, color="7A5A00")
    else:
        variance = round(calculated_total - expected_total, 2)
        if abs(variance) <= 0.01:
            verification.value = f"VERIFIED - Matches the expected total of ${expected_total:,.2f}."
            verification.fill = PatternFill("solid", fgColor=PALE_GREEN)
            verification.font = Font(name="Aptos", size=9, bold=True, color=GREEN)
        elif variance > 0:
            verification.value = (
                f"WARNING - ${variance:,.2f} above the expected ${expected_total:,.2f}. "
                "Check duplicate orders, linehaul/fuel amounts, and accessorial charges."
            )
            verification.fill = PatternFill("solid", fgColor="FCE8E3")
            verification.font = Font(name="Aptos", size=9, bold=True, color="8B3529")
        else:
            verification.value = (
                f"WARNING - ${abs(variance):,.2f} below the expected ${expected_total:,.2f}. "
                "Check excluded or missing orders and missing charges."
            )
            verification.fill = PatternFill("solid", fgColor="FFF5D6")
            verification.font = Font(name="Aptos", size=9, bold=True, color="7A5A00")
    verification.alignment = Alignment(vertical="center")

    sheet.merge_cells("A9:K9")
    sheet["A9"] = "REVIEWED ORDER & WAREHOUSE MOVEMENT BREAKDOWN"
    sheet["A9"].font = Font(name="Aptos", size=9, bold=True, color=GREEN)
    sheet["A9"].alignment = Alignment(vertical="center")

    headers = [
        "Order / movement",
        "Delivery / BOL",
        "Warehouse",
        "City / state / ZIP",
        "Trailer",
        "Miles",
        "Pallets / drop",
        "Linehaul / allocation",
        "Fuel / allocation",
        "Accessorials",
        "Order total",
    ]
    header_row = 10
    first_data_row = 11
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, header)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=GOLD))
    sheet.row_dimensions[header_row].height = 30

    row_number = first_data_row
    for group_index, order_stops in enumerate(grouped.values()):
        first = order_stops[0]
        last = order_stops[-1]
        accessorials = first.extra_drops + first.extra_pickups + first.other_charges
        order_values = [
            f"{first.company_id} · {first.order_id}" if first.company_id else first.order_id,
            first.bol_number,
            (
                f"Route: {_location(first.origin_city, first.origin_state)} → "
                f"{_location(last.destination_city, last.destination_state)}"
            ),
            f"{len(order_stops)} warehouse stop{'s' if len(order_stops) != 1 else ''}",
            first.trailer_number,
            first.miles,
            first.total_pallets,
            first.freight_charge,
            first.fuel_surcharge,
            accessorials,
            first.total_charge,
        ]
        group_fill = PALE_GREEN if group_index % 2 == 0 else PALE_BLUE
        for column, value in enumerate(order_values, start=1):
            cell = sheet.cell(row_number, column, value)
            cell.fill = PatternFill("solid", fgColor=group_fill)
            cell.border = Border(top=Side(style="medium", color="BAC8BD"))
            cell.font = Font(name="Aptos", size=9, bold=True, color=INK)
            cell.alignment = Alignment(vertical="center")
        for column in range(8, 12):
            sheet.cell(row_number, column).number_format = '$#,##0.00'
        sheet.cell(row_number, 6).number_format = '#,##0.0'
        sheet.cell(row_number, 7).number_format = '#,##0'
        sheet.row_dimensions[row_number].height = 21
        row_number += 1

        for stop in order_stops:
            destination = " ".join(
                part
                for part in (
                    _location(stop.destination_city, stop.destination_state),
                    stop.destination_zip.strip(),
                )
                if part
            )
            # Movement rows explain where the load went without duplicating the
            # order-level billing amount shown on the bold parent row.
            movement_values = [
                f"   ↳ Movement {stop.movement_sequence}",
                _excel_date(stop.delivery_date),
                stop.consignee or "Warehouse / delivery stop",
                destination,
                "",
                "",
                stop.pallets_dropped,
                stop.allocated_freight,
                stop.allocated_fuel,
                "",
                "",
            ]
            for column, value in enumerate(movement_values, start=1):
                cell = sheet.cell(row_number, column, value)
                cell.fill = PatternFill("solid", fgColor="F8FAF8")
                cell.border = Border(bottom=Side(style="thin", color=LINE))
                cell.font = Font(name="Aptos", size=8, color=MUTED)
                cell.alignment = Alignment(vertical="center")
            sheet.cell(row_number, 3).font = Font(
                name="Aptos", size=8, bold=True, color=INK
            )
            sheet.cell(row_number, 2).number_format = "yyyy-mm-dd hh:mm"
            sheet.cell(row_number, 7).number_format = '#,##0'
            for column in (8, 9):
                sheet.cell(row_number, column).number_format = '$#,##0.00'
            sheet.row_dimensions[row_number].outlineLevel = 1
            sheet.row_dimensions[row_number].height = 18
            row_number += 1

    last_data_row = row_number - 1
    total_row = last_data_row + 2
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=9)
    sheet.cell(total_row, 1, "VERIFIED INVOICE TOTAL")
    sheet.cell(total_row, 1).font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    sheet.cell(total_row, 1).fill = PatternFill("solid", fgColor=DARK_GREEN)
    sheet.cell(total_row, 1).alignment = Alignment(horizontal="right", vertical="center")
    sheet.merge_cells(start_row=total_row, start_column=10, end_row=total_row, end_column=11)
    total_cell = sheet.cell(total_row, 10)
    total_cell.value = f"=SUM(K{first_data_row}:K{last_data_row})"
    total_cell.font = Font(name="Aptos Display", size=15, bold=True, color=WHITE)
    total_cell.fill = PatternFill("solid", fgColor=DARK_GREEN)
    total_cell.number_format = '$#,##0.00'
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    sheet.row_dimensions[total_row].height = 28

    note_row = total_row + 2
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 1, end_column=11)
    note = sheet.cell(note_row, 1)
    note.value = (
        "Each bold order row carries the order-level charges once. The indented rows show every movement in "
        "sequence, including delivery time, stop, consignee, pallet drop, and any pallet-based allocation. "
        "See Load Detail for the complete source fields."
    )
    note.font = Font(name="Aptos", size=8, italic=True, color=MUTED)
    note.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [19, 20, 35, 27, 14, 11, 13, 17, 15, 15, 15]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.print_title_rows = "1:10"
    sheet.print_area = f"A1:K{note_row + 1}"


def _build_detail_sheet(workbook: Workbook, stops: list[CustomerStop]) -> None:
    """Add the auditable stop-level source sheet behind the invoice summary."""
    sheet = workbook.create_sheet("Load Detail")
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A7"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    _style_title(
        sheet,
        "LOAD DETAIL",
        "Stop-level support for the reviewed customer invoice",
        "AC",
    )

    header_row = 6
    first_data_row = 7
    for column, header in enumerate(DETAIL_HEADERS, start=1):
        cell = sheet.cell(header_row, column, header)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(name="Aptos", size=8, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=GOLD))
    sheet.row_dimensions[header_row].height = 34

    previous_order = None
    current_color = PALE_BLUE
    for row_number, stop in enumerate(stops, start=first_data_row):
        order_key = (stop.company_id, stop.order_id)
        if order_key != previous_order:
            current_color = PALE_GREEN if current_color == PALE_BLUE else PALE_BLUE
        values = [
            stop.order_id,
            _excel_date(stop.ordered_date),
            _excel_date(stop.delivery_date),
            _excel_date(stop.bill_date),
            stop.origin_city,
            stop.origin_state,
            stop.origin_zip,
            stop.destination_city,
            stop.destination_state,
            stop.destination_zip,
            stop.consignee,
            stop.miles,
            stop.bol_number,
            stop.commodity,
            stop.weight,
            stop.movement_sequence,
            stop.total_pallets,
            stop.pallets_dropped,
            stop.pallets_picked_up,
            stop.freight_charge,
            stop.fuel_surcharge,
            stop.extra_drops,
            stop.extra_pickups,
            stop.other_charges,
            stop.other_charge_total,
            stop.total_charge,
            # Formula allocations make the workbook auditable and recalculate if
            # a reviewer changes pallet counts or charge values in Excel.
            f'=IF($Q{row_number}=0,0,$U{row_number}*$R{row_number}/$Q{row_number})',
            f'=IF($Q{row_number}=0,0,$T{row_number}*$R{row_number}/$Q{row_number})',
            stop.trailer_number,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, value)
            cell.fill = PatternFill("solid", fgColor=current_color)
            cell.border = Border(bottom=Side(style="thin", color=LINE))
            cell.font = Font(name="Aptos", size=8, color=INK)
            cell.alignment = Alignment(vertical="center")
        previous_order = order_key

    last_data_row = first_data_row + len(stops) - 1
    for column in range(20, 29):
        for row in range(first_data_row, last_data_row + 1):
            sheet.cell(row, column).number_format = '$#,##0.00'
    for column in (2, 3, 4):
        for row in range(first_data_row, last_data_row + 1):
            sheet.cell(row, column).number_format = "yyyy-mm-dd"

    table = Table(displayName="ReviewedLoadDetail", ref=f"A{header_row}:AC{last_data_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    widths = {
        "A": 14, "B": 12, "C": 12, "D": 12, "E": 16, "F": 9, "G": 11,
        "H": 18, "I": 10, "J": 11, "K": 24, "L": 10, "M": 16, "N": 16,
        "O": 11, "P": 10, "Q": 12, "R": 13, "S": 13, "T": 13, "U": 13,
        "V": 12, "W": 12, "X": 13, "Y": 16, "Z": 14, "AA": 14, "AB": 16, "AC": 15,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.print_title_rows = "1:6"
    sheet.print_area = f"A1:AC{last_data_row}"


def generate_customer_invoice(
    stops: list[CustomerStop],
    invoice_number: str = "",
    bill_date: date | str = "",
    expected_total: float | None = None,
) -> bytes:
    """Create the reviewed workbook and return its XLSX bytes for download."""
    if not stops:
        raise ValueError("No customer loads were found for this bill date")

    calculated_total = invoice_total(stops)
    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    grouped = group_stops_by_order(stops)
    _build_summary_sheet(
        workbook,
        grouped,
        invoice_number,
        bill_date,
        calculated_total,
        None if expected_total is None else round(expected_total, 2),
    )
    _build_detail_sheet(workbook, stops)
    workbook.active = 0

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
